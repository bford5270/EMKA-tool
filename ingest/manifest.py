"""EMKA_Corpus_Manifest.xlsx -> normalized manifest rows.

The manifest is the source of truth for what gets ingested (DESIGN.md §1).
This module READS it; nothing in the codebase ever writes to the xlsx —
ingest status and hashes go to the local ingest log (ingest/ingest_log.py).

Every document sheet shares the same layout: banner rows 1–3, header at
row 4 (`Ref ID | Title | Source / Publisher | Edition / Date | Tier |
Category | Rights | Shelf | Phase | Acquisition Source / Notes |
Status (custodian) | SHA-256 (at ingest)`), data from row 5.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

# Sheets that don't hold document rows.
NON_DOCUMENT_SHEETS = {"README"}

EXPECTED_HEADER_PREFIX = ["Ref ID", "Title"]


@dataclass(frozen=True)
class ManifestRow:
    ref_id: str
    title: str
    source: str
    edition_date: str
    tier: str  # normalized: T1..T5
    category: str
    rights: str  # e.g. "USG", "©-LIC", "VERIFY"
    shelf: str  # normalized: authoritative | unit | deferred
    phase: str  # normalized: "0", "1", "2", ...
    acquisition: str
    status: str  # custodian status as written in the manifest
    sheet: str

    @property
    def rights_gated(self) -> bool:
        """©-licensed or unverified-rights rows are listed, never ingested,
        until the custodian clears them."""
        r = self.rights.upper()
        return "©" in self.rights or "LIC" in r or "VERIFY" in r

    @property
    def is_phase_core(self) -> bool:
        return self.phase == "0"


def _normalize_shelf(raw: str) -> str:
    s = (raw or "").strip().lower()
    if s.startswith("authoritative"):
        return "authoritative"
    if s.startswith("unit"):
        return "unit"
    return "deferred"


def _normalize_phase(raw: object) -> str:
    # "0 — Core" (em dash), "0 - Core", plain "1"/"2", or numeric cells
    s = str(raw or "").strip()
    return s.split()[0].split("—")[0].split("-")[0].strip() if s else ""


def _normalize_tier(raw: str) -> str:
    s = (raw or "").strip().upper()
    return s if s in {"T1", "T2", "T3", "T4", "T5"} else "T5"


def load_manifest(xlsx_path: str | Path) -> list[ManifestRow]:
    import openpyxl

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    rows: list[ManifestRow] = []
    for name in wb.sheetnames:
        if name in NON_DOCUMENT_SHEETS:
            continue
        ws = wb[name]
        header_seen = False
        for values in ws.iter_rows(values_only=True):
            cells = ["" if v is None else str(v).strip() for v in values]
            cells += [""] * (12 - len(cells))
            if not header_seen:
                if cells[:2] == EXPECTED_HEADER_PREFIX:
                    header_seen = True
                continue
            if not cells[0]:
                continue
            rows.append(
                ManifestRow(
                    ref_id=cells[0],
                    title=cells[1],
                    source=cells[2],
                    edition_date=cells[3],
                    tier=_normalize_tier(cells[4]),
                    category=cells[5],
                    rights=cells[6],
                    shelf=_normalize_shelf(cells[7]),
                    phase=_normalize_phase(cells[8]),
                    acquisition=cells[9],
                    status=cells[10],
                    sheet=name,
                )
            )
    wb.close()
    return rows
